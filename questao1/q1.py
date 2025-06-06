import csv
import re
from unidecode import unidecode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def limpar_cpf(cpf):
    return re.sub(r'\D', '', cpf).zfill(11)

def ajustar_largura(pagina):
    for col in pagina.columns:
        tam = max(len(str(cell.value or "")) for cell in col)
        pagina.column_dimensions[get_column_letter(col[0].column)].width = tam + 3

arquivo_txt = "dados.txt"
arquivo_excel = "dados_formatados.xlsx"

# Usando openpyxl para criar e manipular o arquivo Excel
# Criando um novo arquivo Excel
planilha = Workbook()
pagina = planilha.active

# Ler e processar o arquivo
with open(arquivo_txt, encoding='utf-8') as f:
    for i, linha in enumerate(csv.reader(f, delimiter='|')):
        valores = []
        for j, valor in enumerate(linha):
            valor = unidecode(valor.strip()).upper()
            if i > 0 and j == 1:
                valor = limpar_cpf(valor)
            valores.append(valor)
        pagina.append(valores)

# Aplicar formatações
pagina.auto_filter.ref = pagina.dimensions
pagina.freeze_panes = "A2"
ajustar_largura(pagina)

# Salvar
planilha.save(arquivo_excel)



